import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment

def generate_report(file_path, ai_analysis):
    df = pd.read_csv(file_path)
    wb = Workbook()
    
    # ===== Sheet 1: 原始数据 =====
    ws1 = wb.active
    ws1.title = "原始数据"
    
    # 写入表头
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws1.cell(row=row_idx+2, column=col_idx, value=value)
    
    # ===== Sheet 2: 统计摘要 =====
    ws2 = wb.create_sheet("统计摘要")
    summary = df.groupby('产品类别')['销售额'].agg(['sum', 'mean', 'count'])
    summary.columns = ['总销售额', '平均销售额', '订单数']
    summary = summary.reset_index()
    
    for col, header in enumerate(summary.columns, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row in summary.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws2.cell(row=row_idx+2, column=col_idx, value=value)
    
    # ===== Sheet 3: AI分析结果 =====
    ws3 = wb.create_sheet("AI分析")
    ws3['A1'] = "AI 数据分析报告"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A2'] = ai_analysis
    ws3['A2'].alignment = Alignment(wrap_text=True)
    ws3.column_dimensions['A'].width = 80
    
    # ===== Sheet 4: 图表 =====
    ws4 = wb.create_sheet("可视化图表")
    img = Image("D:\\mycode\\sales_report.png")
    img.width = 700
    img.height = 500
    ws4.add_image(img, "A1")
    
    output_path = "D:\\mycode\\sales_analysis_report.xlsx"
    wb.save(output_path)
    print(f"Excel报告已生成：{output_path}")

if __name__ == "__main__":
    import requests
    
    df = pd.read_csv("D:\\mycode\\sales_data.csv")
    
    prompt = f"""
    我有一份数据集，基本信息如下：
    列名：{list(df.columns)}
    行数：{df.shape[0]}
    基本统计：{df.describe().to_string()}
    请用中文分析这份数据，指出有趣的规律和洞察。
    """
    
    print("正在让AI分析数据，请稍等...")
    response = requests.post(
        "http://localhost:11434/api/generate",
        json={"model": "gemma4:e4b", "prompt": prompt, "stream": False}
    )
    ai_analysis = response.json().get("response", "无法获取回复")
    print("AI分析完成！正在生成报告...")
    
    generate_report("D:\\mycode\\sales_data.csv", ai_analysis)